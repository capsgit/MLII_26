from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# =========================
# CONFIG
# =========================

LEARNING_RATE: float = 0.1
BIAS: float = 0.2
INITIAL_W1: float = 0.3
INITIAL_W2: float = -0.1

AND_MATRIX = [
    [0, 0, 0],
    [0, 1, 0],
    [1, 0, 0],
    [1, 1, 1],
]

OR_MATRIX = [
    [0, 0, 0],
    [0, 1, 1],
    [1, 0, 1],
    [1, 1, 1],
]


# =========================
# PERCEPTRON LOGIC
# =========================

def step(net: float) -> int:
    """
    Binary step activation function.

    Returns
    -------
    int
        1 if net >= 0, otherwise 0.
    """
    return 1 if net >= 0 else 0


def net_input(x1: int, x2: int, w1: float, w2: float, b: float) -> float:
    """
    Compute the weighted sum minus bias.
    """
    return (x1 * w1 + x2 * w2) - b


def calculate_error(y_desired: int, y_actual: int) -> int:
    """
    Compute the perceptron error.
    """
    return y_desired - y_actual


def train_perceptron(
    training_data: list[list[int]],
    w1: float,
    w2: float,
    b: float,
    learning_rate: float,
    max_epochs: int = 100,
) -> list[dict]:
    """
    Train a simple perceptron until convergence.

    A new epoch is created only if there is at least one non-zero error
    in the current epoch. Training stops when all errors in one full epoch
    are zero, or when max_epochs is reached.

    Parameters
    ----------
    training_data : list[list[int]]
        Rows in the form [x1, x2, y_desired].
    w1 : float
        Initial weight for input x1.
    w2 : float
        Initial weight for input x2.
    b : float
        Bias term (kept constant in this implementation).
    learning_rate : float
        Learning rate alpha.
    max_epochs : int, default=100
        Safety limit to avoid infinite loops.

    Returns
    -------
    list[dict]
        Training history with one record per processed row.
    """
    history: list[dict] = []
    epoch = 1

    while epoch <= max_epochs:
        print(f"\n--- Epoch {epoch} ---")
        had_error = False

        for x1, x2, y_desired in training_data:
            w1_before = w1
            w2_before = w2

            net = net_input(x1, x2, w1_before, w2_before, b)
            y_actual = step(net)
            error = calculate_error(y_desired, y_actual)

            if error != 0:
                had_error = True

            delta_w1 = learning_rate * x1 * error
            delta_w2 = learning_rate * x2 * error

            w1_after = w1_before + delta_w1
            w2_after = w2_before + delta_w2

            row = {
                "epoch": epoch,
                "x1": x1,
                "x2": x2,
                "y_desired": y_desired,
                "net": net,
                "y_actual": y_actual,
                "error": error,
                "delta_w1": delta_w1,
                "delta_w2": delta_w2,
                "w1_before": w1_before,
                "w2_before": w2_before,
                "w1_after": w1_after,
                "w2_after": w2_after,
            }
            history.append(row)

            print(
                f"x1={x1}, x2={x2}, yd={y_desired}, "
                f"net={net:.1f}, ya={y_actual}, e={error}, "
                f"Δw1={delta_w1:.1f}, Δw2={delta_w2:.1f}, "
                f"w1_before={w1_before:.1f}, w2_before={w2_before:.1f}, "
                f"w1_after={w1_after:.1f}, w2_after={w2_after:.1f}"
            )

            w1 = w1_after
            w2 = w2_after

        if not had_error:
            print(f"\nConverged at epoch {epoch}: all errors were 0.")
            break

        epoch += 1
    else:
        print(f"\nStopped at max_epochs={max_epochs} without full convergence.")

    return history


# =========================
# EXCEL STYLES
# =========================

thin = Side(style="thin", color="000000")
BORDER_ALL = Border(left=thin, right=thin, top=thin, bottom=thin)

FILL_BLUE = PatternFill("solid", fgColor="5B9BD5")
FILL_LIGHT = PatternFill("solid", fgColor="D9E2F3")
FILL_ERROR = PatternFill("solid", fgColor="FDE9E7")
FILL_OK = PatternFill("solid", fgColor="E2F0D9")

FONT_TITLE = Font(size=16, bold=True)
FONT_HEADER_WHITE = Font(bold=True, color="FFFFFF")
FONT_SUBHEADER = Font(bold=True)
FONT_RED = Font(bold=True, color="FF0000")
FONT_GREEN = Font(bold=True, color="008000")
FONT_PURPLE = Font(bold=True, color="7030A0")
FONT_SOFT_PINK = Font(bold=True, color="C27BA0")
FONT_BLACK = Font(bold=True, color="000000")

CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def style_cell(cell, fill=None, font=None, alignment=CENTER, border=True) -> None:
    if fill is not None:
        cell.fill = fill
    if font is not None:
        cell.font = font
    if border:
        cell.border = BORDER_ALL
    cell.alignment = alignment


def apply_border_range(ws, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    for row in ws.iter_rows(
        min_row=start_row,
        max_row=end_row,
        min_col=start_col,
        max_col=end_col,
    ):
        for cell in row:
            cell.border = BORDER_ALL
            cell.alignment = CENTER


# =========================
# EXCEL WRITING HELPERS
# =========================

def format_decimal(value: float) -> float:
    """
    Keep one decimal place visually consistent in Excel.
    """
    return round(value, 1)


def write_title(ws, gate_name: str) -> None:
    ws.merge_cells("A1:F1")
    cell = ws["A1"]
    cell.value = f"{gate_name} — perceptron training history"
    cell.font = FONT_TITLE
    cell.alignment = LEFT


def write_parameters_box(ws, start_row: int, start_col: int, params: dict) -> None:
    ws.merge_cells(
        start_row=start_row,
        start_column=start_col,
        end_row=start_row,
        end_column=start_col + 1,
    )
    header = ws.cell(start_row, start_col, "Parameters")
    style_cell(header, fill=FILL_BLUE, font=FONT_HEADER_WHITE)
    apply_border_range(ws, start_row, start_row, start_col, start_col + 1)

    rows = [
        ("learning_rate (alpha)", params["learning_rate"]),
        ("bias (b)", params["bias"]),
        ("initial_w1", params["initial_w1"]),
        ("initial_w2", params["initial_w2"]),
        ("step rule", params["step_rule"]),
    ]

    row_idx = start_row + 1
    for key, value in rows:
        c1 = ws.cell(row_idx, start_col, key)
        c2 = ws.cell(row_idx, start_col + 1, value)
        style_cell(c1, fill=FILL_LIGHT, font=FONT_SUBHEADER, alignment=LEFT)
        style_cell(c2, alignment=CENTER)
        row_idx += 1


def write_truth_table_box(
    ws,
    start_row: int,
    start_col: int,
    truth_table: list[list[int]],
    gate_name: str,
) -> None:
    ws.merge_cells(
        start_row=start_row,
        start_column=start_col,
        end_row=start_row,
        end_column=start_col + 2,
    )
    header = ws.cell(start_row, start_col, f"{gate_name} truth table")
    style_cell(header, fill=FILL_BLUE, font=FONT_HEADER_WHITE)
    apply_border_range(ws, start_row, start_row, start_col, start_col + 2)

    headers = ["x1", "x2", "desired"]
    for offset, value in enumerate(headers):
        cell = ws.cell(start_row + 1, start_col + offset, value)
        style_cell(cell, fill=FILL_LIGHT, font=FONT_SUBHEADER)

    row_idx = start_row + 2
    for x1, x2, yd in truth_table:
        for col_idx, value in enumerate((x1, x2, yd), start=start_col):
            cell = ws.cell(row_idx, col_idx, value)
            style_cell(cell)
        row_idx += 1


def write_main_header(ws, row: int) -> None:
    headers = [
        "Epoch",
        "x1",
        "x2",
        "y_d",
        "net",
        "y_a",
        "e",
        "Δw1",
        "Δw2",
        "w1_before",
        "w2_before",
        "w1_after",
        "w2_after",
    ]

    for col_idx, header_text in enumerate(headers, start=1):
        cell = ws.cell(row, col_idx, header_text)
        style_cell(cell, fill=FILL_BLUE, font=FONT_HEADER_WHITE)

        if header_text == "y_a":
            cell.font = FONT_SOFT_PINK
        elif header_text == "e":
            cell.font = FONT_PURPLE
        elif header_text in ("w1_after", "w2_after"):
            cell.font = FONT_RED


def group_history_by_epoch(history: list[dict]) -> dict[int, list[dict]]:
    grouped: dict[int, list[dict]] = {}
    for row in history:
        grouped.setdefault(row["epoch"], []).append(row)
    return grouped


def write_epoch_block(ws, start_row: int, epoch_number: int, epoch_rows: list[dict]) -> int:
    ws.merge_cells(
        start_row=start_row,
        start_column=1,
        end_row=start_row,
        end_column=13,
    )
    epoch_cell = ws.cell(start_row, 1, f"Epoch #{epoch_number}")
    style_cell(epoch_cell, fill=FILL_BLUE, font=FONT_HEADER_WHITE, alignment=LEFT)
    apply_border_range(ws, start_row, start_row, 1, 13)

    current_row = start_row + 1

    for row_data in epoch_rows:
        values = [
            row_data["epoch"],
            row_data["x1"],
            row_data["x2"],
            row_data["y_desired"],
            format_decimal(row_data["net"]),
            row_data["y_actual"],
            row_data["error"],
            format_decimal(row_data["delta_w1"]),
            format_decimal(row_data["delta_w2"]),
            format_decimal(row_data["w1_before"]),
            format_decimal(row_data["w2_before"]),
            format_decimal(row_data["w1_after"]),
            format_decimal(row_data["w2_after"]),
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(current_row, col_idx, value)
            style_cell(cell)

            if col_idx in (5, 8, 9, 10, 11, 12, 13):
                cell.number_format = "0.0"

            if col_idx == 6:  # y_a
                cell.font = FONT_SOFT_PINK
            elif col_idx == 7:  # error
                cell.font = FONT_PURPLE
                cell.fill = FILL_OK if value == 0 else FILL_ERROR
            elif col_idx in (8, 9):  # deltas
                if value != 0:
                    cell.font = FONT_GREEN
            elif col_idx in (12, 13):  # after
                cell.font = FONT_RED

        current_row += 1

    return current_row + 1


def set_column_widths(ws) -> None:
    widths = {
        1: 10,
        2: 8,
        3: 8,
        4: 8,
        5: 10,
        6: 8,
        7: 8,
        8: 8,
        9: 8,
        10: 12,
        11: 12,
        12: 11,
        13: 11,
    }
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_training_sheet(
    ws,
    gate_name: str,
    history: list[dict],
    truth_table: list[list[int]],
    params: dict,
) -> None:
    write_title(ws, gate_name)
    write_parameters_box(ws, start_row=3, start_col=2, params=params)
    write_truth_table_box(ws, start_row=3, start_col=6, truth_table=truth_table, gate_name=gate_name)
    write_main_header(ws, row=10)

    grouped = group_history_by_epoch(history)
    next_row = 11

    for epoch_number in sorted(grouped.keys()):
        next_row = write_epoch_block(ws, next_row, epoch_number, grouped[epoch_number])

    set_column_widths(ws)
    ws.freeze_panes = "A10"


def write_summary_sheet(ws, and_history: list[dict], or_history: list[dict]) -> None:
    ws["A1"] = "Summary"
    ws["A1"].font = FONT_TITLE

    headers = ["Gate", "Epochs used", "Final w1", "Final w2", "Converged"]
    for col_idx, text in enumerate(headers, start=1):
        cell = ws.cell(3, col_idx, text)
        style_cell(cell, fill=FILL_BLUE, font=FONT_HEADER_WHITE)

    def summarize(history: list[dict], gate_name: str, row_idx: int) -> None:
        last = history[-1]
        epochs_used = max(item["epoch"] for item in history)
        values = [
            gate_name,
            epochs_used,
            format_decimal(last["w1_after"]),
            format_decimal(last["w2_after"]),
            "yes" if all(item["error"] == 0 for item in history[-4:]) else "check",
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row_idx, col_idx, value)
            style_cell(cell)

    summarize(and_history, "AND", 4)
    summarize(or_history, "OR", 5)

    for col_idx, width in {1: 12, 2: 12, 3: 12, 4: 12, 5: 12}.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def create_workbook(and_history: list[dict], or_history: list[dict], output_path: str) -> None:
    wb = Workbook()

    params = {
        "learning_rate": LEARNING_RATE,
        "bias": BIAS,
        "initial_w1": INITIAL_W1,
        "initial_w2": INITIAL_W2,
        "step_rule": "1 if net >= 0 else 0",
    }

    ws_and = wb.active
    ws_and.title = "AND_Gate"
    write_training_sheet(ws_and, "AND Gate", and_history, AND_MATRIX, params)

    ws_or = wb.create_sheet("OR_Gate")
    write_training_sheet(ws_or, "OR Gate", or_history, OR_MATRIX, params)

    ws_summary = wb.create_sheet("Summary")
    write_summary_sheet(ws_summary, and_history, or_history)

    wb.save(output_path)


def main() -> None:
    print("TRAINING AND GATE")
    and_history = train_perceptron(
        training_data=AND_MATRIX,
        w1=INITIAL_W1,
        w2=INITIAL_W2,
        b=BIAS,
        learning_rate=LEARNING_RATE,
        max_epochs=100,
    )

    print("\n" + "=" * 60)
    print("TRAINING OR GATE")
    or_history = train_perceptron(
        training_data=OR_MATRIX,
        w1=INITIAL_W1,
        w2=INITIAL_W2,
        b=BIAS,
        learning_rate=LEARNING_RATE,
        max_epochs=100,
    )

    output_file = "perceptron_gates_full.xlsx"
    create_workbook(and_history, or_history, output_file)

    print("\nWorkbook created successfully:")
    print(output_file)


if __name__ == "__main__":
    main()
